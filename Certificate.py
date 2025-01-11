import io
from openpyxl.reader.excel import load_workbook
from reportlab.lib import pagesizes
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas
import EmbedOnPDF
import re


def nameFormatter(fullName):
    modName = ""
    modList = re.split('[. -]', str(fullName))
    initialAdded = False
    try:
        for i in modList:
            if i != "" and i.isalpha():
                if len(i) == 1:
                    initialAdded = True
                    modName = modName + i.upper()
                elif len(i) == 2:
                    initialAdded = True
                    modName = modName + i.upper()
                else:
                    if initialAdded:
                        modName = modName + " " + i.capitalize()
                    else:
                        modName = modName + i.capitalize() + " "
        if modName[-1] == " ":
            modName = modName[:len(modName) - 1]
    except IndexError:
        modName = ""
    return modName


def create_certificates(dataFile, templatePdf, certificatesFolder, pdfOnly):
    wb = load_workbook(filename=dataFile, data_only=True)
    sheetData = wb.sheetnames
    for sheet_name in sheetData:
        data = wb[sheet_name]
        for row_num, row in enumerate(data):
            if row[0].value is None:
                break
            if row_num == 0 or row_num == 1:
                continue
            rollNo = str(int(row[2].value))
            name = nameFormatter(row[3].value)
            school = row[4].value.upper()
            date = "2025 January 12, 10 AM"
            rowlist = [school, sheet_name, name, date]
            packet = io.BytesIO()
            pdfmetrics.registerFont(TTFont('GothamMedium', 'assets/GothamMedium.ttf'))
            pdfmetrics.registerFont(TTFont('GothamLight', 'assets/GothamLight.ttf'))
            pdfmetrics.registerFont(TTFont('GothamBook', 'assets/GothamBook.ttf'))
            can = canvas.Canvas(packet, pagesize=pagesizes.landscape(pagesizes.B0))
            can.setFont('GothamMedium', 14)
            can.setFillColor("#ef5362")
            can.drawString(344, 474.5, rollNo)
            can.setFont('GothamLight', 10.5)
            can.setFillColor("#000000")
            for i in range(4):
                can.drawString(219, 355 + i * 21, rowlist[i])
            print(sheet_name, name, school)
            can.save()
            # move to the beginning of the StringIO buffer
            packet.seek(0)
            EmbedOnPDF.embedOnPDF(certificatesFolder, sheet_name, name, packet, templatePdf, pdfOnly)


# Replace these with your file paths and starting number
data_file = "sheet.xlsx"
template_pdf = "certificate.pdf"
certificates_folder = "Certificates"
saveState = 0  # 0: Only Pdf, 1: Only Image, 2: Both Pdf and Image
create_certificates(data_file, template_pdf, certificates_folder, saveState)
print("Certificates created successfully!")
